import os

# For next imports you will need to install 5 packages:
# pip install openpyxl
# pip install pandas
# pip install pdfminer
# pip install pdfminer.six
# pip install xlsxwriter

import openpyxl 
import pandas as pd 
from pdfminer.high_level import extract_pages
from pdfminer.layout import LTTextContainer

"""
excel functions:
load keywords from A column
set new excel table column sizes
"""

def load_keys(xlsx_file: str):
    wb = openpyxl.load_workbook(xlsx_file)  # load excel table
    ws = wb.active
    # get all values from A column (keywords)
    keys = list(dict.fromkeys([str(cell.value) for cell in ws["A"]]))
    wb.close()
    return keys

def set_size(xlsx_file: str):
    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb.worksheets[0]

    def as_text(value):
        if value is None:
            return ""
        return str(value)

    for column_cells in ws.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[openpyxl.utils.get_column_letter(
            column_cells[0].column)].width = length
    # set fixed width for column A and B
    ws.column_dimensions['A'].width = 65
    ws.column_dimensions['B'].width = 65
    wb.save(xlsx_file)
    wb.close()

"""
pdf functions:
read a pdf file
format read text
search for a keywoard in the text
"""

def get_text(file):
    text = ""
    for page in extract_pages(file):
        for element in page:
            if isinstance(element, LTTextContainer): # element is text
                text += element.get_text() 
        text += "\n"
    return text

special_symbols = [".", "\n", ":", "-", ";", ",", "!", "?", "^"]

def prepare_text(text: str) -> tuple:
    text = text.lower()
    for symbol in special_symbols:
        text = text.replace(symbol, " ")
    words = text.split()
    text = ""
    for word in words:
        text += word + " "
    return (text, words)
    

def search_keywords(keyword: str, text_res: tuple):
    for symbol in special_symbols:
        keyword = keyword.lower().replace(symbol, " ")

    text_str = text_res[0]
    text = text_res[1]
    if keyword.count(' ') == 0:
        return text.count(keyword)
    else:
        return text_str.count(keyword)

    
    

"""
pdf files handling functions:
load all pdf files as a list of paths
"""
def load_pdfs():
    pdf_files = []
    while (path := input("Path to pdf file/directory ([-] to start analyze): ")) != "-":
        if os.path.isfile(path) and path.endswith(".pdf"):
            pdf_files.append(path)
        elif os.path.isdir(path):
            # loop over all files and find only pdf
            pdfs = [os.path.join(path, i) for i in os.listdir(path) if i.endswith(".pdf")]  
            pdf_files += pdfs
    return list(dict.fromkeys(pdf_files))

def process_pdf(file):
    print(f"Processing file: {file}")
    short_file = file.split("\\")[-1]
    keywords_found = []
    text_data = prepare_text(get_text(file))

    for key in keywords:  # search keywords in every file
        print(f"Searching keyword: {key}")  
        result = search_keywords(key, text_data)
        if result > 0:
            keywords_found.append(key)
            file_key_count.append([short_file, key, result])
        if data.get(short_file, None) != None:  # save results to dictionary
            data[short_file].append(result)
        else:
            data[short_file] = [result]
        if result == 0:
            file_key_count.append([short_file, key, result])                

    data[short_file].append("; ".join(keywords_found))
    all_keywords.extend(keywords_found)

"""
main script body
"""

version = "1.4.0"
print(f"Developed by xymoget. Version {version}\nxymoget@gmail.com")

keys_path = input("Path to .xlsx file with keywords: ")
save_path = input("Path to .xlsx file to save analysis: ")

pdf_files = load_pdfs()
keywords = load_keys(keys_path)

data = {
    "keyword": keywords
}

all_keywords = []
file_key_count = []

for file in pdf_files:  # loop over all pdf files provided
    process_pdf(file)

data["keyword"].append("Keywords found")
data["keyword"].append("Total")
data["keyword"].append("Total per single keyword")
data["conclusion"] = []
print("Summing up...")

# find files where keywords were used
for i in range(len(data["keyword"]) - 3):
    counter = 0
    files = []
    for key in data.keys():
        if key == "keyword" or key == "conclusion":
            continue
        if data[key][i] > 0:
            counter += 1
            files.append(f'"{key}"')
    conclusion = f"{str(counter)} file(s): {'; '.join(files)}"
    data["conclusion"].append(conclusion)

for key in data.keys():  # calculate total amount of keywords used in file
    if key == "keyword" or key == "conclusion":
        continue
    data[key].append(sum(data[key][:-1]))
    data[key].append(len([i for i in data[key][:-2] if i > 0]))

unique_keywords = list(set(all_keywords))
data["conclusion"] += [f'{len(unique_keywords)} keywords found: {"; ".join(unique_keywords)}', sum([data[i][-2]
                                                                                                    for i in data if i not in ["keyword", "conclusion"]]), sum([data[i][-1] for i in data if i not in ["keyword", "conclusion"]])]

for key in data:
    temp = data[key][-3]
    data[key][-3] = data[key][-1]
    data[key][-1] = temp

df = pd.DataFrame(data)  # turn dictionary into dataframe object

# this is a list of all rows need to be styled as bold and centered.
styled_rows = []

# first ranking.
def keywords_found_ranking():
    # creating a dict
    file_counts = {}

    # count each keyword presence in files.
    for filename, key, count in file_key_count:
        # if keyword is met more than 0 times in a pdf, k = 1.
        k = 1 if count > 0 else 0
        # adding info to the dict (file_counts)
        if filename in file_counts:
            file_counts[filename] += k
        else:
            file_counts[filename] = k

    # create new dict: count of keywords serves as a key, and the value is filename.
    grouped_files = {}
    for file, count in file_counts.items():
        if count in grouped_files:
            grouped_files[count].append(file)
        else:
            grouped_files[count] = [file]

    # sort it in descending order.
    sorted_counts = sorted(grouped_files.keys(), reverse=True)

    # create lists for each column A and B
    column_A = []
    column_B = []

    # calculate the start of column's contents. (add empty rows to the end)
    startrow = len(df.index)+1 + 3

    # "i" is used to indicate which row we are on. Incrementing it every time a new row (line) is created in the table.
    i = 0
    for count in sorted_counts:

        # column A
        # append data to the A column. each result is in different line.
        files_with_count = grouped_files[count]
        column_A.append(f"PDF file(s) ({len(files_with_count)}) with {count} keyword(s):")
        i += 1
        styled_rows.append(startrow+i)
        for file in files_with_count:
            column_A.append(f"{file}")
            # increment current line after each added file.
            i += 1

        # column B
        # append data to the B column. result is one line, separated by the semicolon.
        column_B.append(f"PDF file(s) ({len(files_with_count)}) with {count} keyword(s):")
        column_B.append("; ".join([f"{file}" for file in files_with_count]))
        for _ in range(len(column_A)-len(column_B)):
            # adding blank spaces to remaining lines, because column B has to be the same length as column A.
            column_B.append('')

    return column_A, column_B


# get a result and turn it into dataframe with 2 columns (A and B).
result = keywords_found_ranking()
keywords_found_ranking = {
    'PDF files ranking by number of keywords found': result[0],
    'PDF files ranking by number of keywords found ': result[1],
}
keywords_found_rankings_df = pd.DataFrame(keywords_found_ranking)

# second ranking
def frequency_ranking() -> list:
    # creating a dict
    file_keywords = {}

    # create entry in a dict, if not created yet. filename serves as a key, keywordcount serves as value. adding keyword count to the key
    for file, key, keyword_count in file_key_count:
        if file in file_keywords:
            file_keywords[file] += keyword_count
        else:
            file_keywords[file] = keyword_count

    # switch key and value, so we can sort it.
    sorted_files = {}
    for file, keyword_count in file_keywords.items():
        if keyword_count in sorted_files:
            sorted_files[keyword_count].append(file)
        else:
            sorted_files[keyword_count] = [file]

    column_A = []
    column_B = []

    # calculate first line of columns's content by adding length of dataframe and length of first ranking, as well as a number of empty rows between them.
    startrow = len(df.index)+1+len(keywords_found_rankings_df.index)+3*2

    i = 0
    for keyword_count, files in sorted(sorted_files.items(), reverse=True):

        # column A
        column_A.append(f"PDF file(s) ({len(files)}) with frequency of {keyword_count}:")
        i += 1
        styled_rows.append(startrow+i)
        for item in files:
            column_A.append(item)
            i += 1

        # column B
        column_B.append(f"PDF file(s) ({len(files)}) with frequency of {keyword_count}:")
        column_B.append(f"{'; '.join(files)}")
        for _ in range(len(column_A)-len(column_B)):
            column_B.append('')

    return column_A, column_B


result = frequency_ranking()
frequency_rankings = {
    'PDF files ranking by frequency/occurrence of keywords appereance': result[0],
    'PDF files ranking by frequency/occurrence of keywords appereance ': result[1]
}
frequency_rankings_df = pd.DataFrame(frequency_rankings)


# third ranking
def ranking_by_keyword_occurence() -> list:

    # create a dict
    keyword_count = {}

    # similarly with the first ranking, but this time we count presence of a keyword, no matter from which file.
    for file, key, count in file_key_count:
        k = 1 if count > 0 else 0
        if key in keyword_count:
            keyword_count[key] += k
        else:
            keyword_count[key] = k

    # sort it by second value
    sorted_keywords = sorted(keyword_count.items(),
                             key=lambda x: x[1], reverse=True)

    sorted_keyword_count = {}
    for keyword, count in sorted_keywords:
        if count in sorted_keyword_count:
            sorted_keyword_count[count].append(keyword)
        else:
            sorted_keyword_count[count] = [keyword]

    column_A = []
    column_B = []

    # similarly to the freq ranking, calculate startrow using lengths of dataframes and empty rows between them.
    startrow = len(df.index)+1 + len(keywords_found_rankings_df.index) + \
        len(frequency_rankings_df.index)+3*3

    i = 0
    for count, keywords in sorted_keyword_count.items():

        # column A
        column_A.append(f"Keyword(s) ({len(keywords)}) found across {count} PDF file(s):")
        i += 1
        styled_rows.append(startrow+i)
        for item in keywords:
            column_A.append(item)
            i += 1

        # column B
        column_B.append(f"Keyword(s) ({len(keywords)}) found across {count} PDF file(s):")
        column_B.append(f"{'; '.join(keywords)}")
        for _ in range(len(column_A)-len(column_B)):
            column_B.append('')

    return column_A, column_B


result = ranking_by_keyword_occurence()
occurence_rankings = {
    'Keywords ranking by number of PDF(s) match': result[0],
    'Keywords ranking by number of PDF(s) match ': result[1]}
keyword_occurence_rankings_df = pd.DataFrame(occurence_rankings)

# fourth ranking
def ranking_keywords_by_freq():

    keyword_counts = {}

    # this time we count not the presence of a keyword, but how many times it is met.
    for file, key, count in file_key_count:
        if key in keyword_counts:
            keyword_counts[key] += count
        else:
            keyword_counts[key] = count

    # combine the values and turn them into list, then sort it.
    unique_counts = list(set(keyword_counts.values()))
    unique_counts.sort(reverse=True)

    column_A = []
    column_B = []

    startrow = len(df.index)+1+len(keywords_found_rankings_df.index) + \
        len(frequency_rankings_df.index) + \
        len(keyword_occurence_rankings_df.index)+3*4

    i = 0
    for count in unique_counts:

        # column A
        # Create a list of keywords with a specific count from the keyword_counts dictionary
        keywords_with_count = [
            keyword for keyword, keyword_count in keyword_counts.items() if keyword_count == count]
        column_A.append(
            f"Keyword(s) ({len(keywords_with_count)}) with frequency value of {count}:")
        i += 1
        styled_rows.append(startrow+i)
        for item in keywords_with_count:
            column_A.append(item)
            i += 1

        # column B
        column_B.append(
            f"Keyword(s) ({len(keywords_with_count)}) with frequency value of {count}:")
        keywords_with_count = '; '.join(
            [keyword for keyword, keyword_count in keyword_counts.items() if keyword_count == count])
        column_B.append(keywords_with_count)
        for _ in range((len(column_A)-len(column_B))):
            column_B.append('')

    return column_A, column_B


result = ranking_keywords_by_freq()
keyword_freq_ranking = {
    'Keywords ranking by frequency/occurrence of appereance': result[0],
    'Keywords ranking by frequency/occurrence of appereance ': result[1]
}
keyword_frequency_rankings_df = pd.DataFrame(keyword_freq_ranking)


with pd.ExcelWriter(save_path, engine='xlsxwriter') as writer:
    # calculating startrow for each dataframe
    first_ranking_startrow = len(df.index) + 3
    second_ranking_startrow = len(
        keywords_found_rankings_df.index) + first_ranking_startrow + 3
    third_ranking_startrow = len(
        frequency_rankings_df.index) + second_ranking_startrow + 3
    fourth_ranking_startrow = len(
        keyword_occurence_rankings_df.index) + third_ranking_startrow + 3

    df.to_excel(writer, sheet_name='Sheet1', index=False)
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    # export dataframes to an excel file
    keywords_found_rankings_df.to_excel(
        writer, sheet_name='Sheet1', index=False, startrow=first_ranking_startrow)
    frequency_rankings_df.to_excel(
        writer, sheet_name='Sheet1', index=False, startrow=second_ranking_startrow)
    keyword_occurence_rankings_df.to_excel(
        writer, sheet_name='Sheet1', index=False, startrow=third_ranking_startrow)
    keyword_frequency_rankings_df.to_excel(
        writer, sheet_name='Sheet1', index=False, startrow=fourth_ranking_startrow)

    # left aligned style
    left_aligned = workbook.add_format({'align': 'left'})
    # apply style to the last column in first dataframe.
    worksheet.set_column(
        len(df.columns)-2, len(df.columns) - 1, None, left_aligned)

    # bold centered style
    bold_centered = workbook.add_format({'bold': True, 'align': 'center'})
    # apply style for each line that has to be styled as bold and centered
    for item in styled_rows:
        worksheet.set_row(item - 1, None, bold_centered)

set_size(save_path)

print(f"Results are saved to {save_path}")
print(df)
print(f"Check {save_path} to see full results (press enter to finish the program)")
input()
